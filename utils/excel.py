import openpyxl
import logging

def clear_excel_sheet(file_path, start_row=7):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        max_row = ws.max_row
        for row in range(start_row, max_row + 1):
            for col in range(2, 6):  # B a E
                ws.cell(row=row, column=col).value = None
        wb.save(file_path)
        return wb, ws
    except Exception as e:
        logging.error(f"Error al limpiar archivo Excel: {e}")
        return None, None

def write_to_excel(data, wb, ws, file_path, start_row=7):
    for i, (lead_id, url, cliente, telefono) in enumerate(data):
        ws.cell(row=start_row + i, column=2, value=lead_id)    # B
        ws.cell(row=start_row + i, column=3, value=url)        # C
        ws.cell(row=start_row + i, column=4, value=cliente)    # D
        ws.cell(row=start_row + i, column=5, value=telefono)   # E
    wb.save(file_path)
    logging.info(f"{len(data)} leads escritos en {file_path}")